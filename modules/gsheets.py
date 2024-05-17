import io
import os
import gspread
import pandas as pd

from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.http import MediaFileUpload
from google.oauth2.service_account import Credentials
from gspread_dataframe import set_with_dataframe
from openpyxl.utils.cell import get_column_letter

from googleapiclient.discovery import build
from ingestaweb.modules.utils import remove_file
from ingestaweb.settings import Config


class GoogleDrive:

    SCOPES = [
        # 'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive',
        # 'https://www.googleapis.com/auth/drive.file'
    ]

    def __init__(self):
        self.creds = self.get_credentials()
        self.gsheets_client = gspread.service_account(
            filename=Config.CREDS_GSHEETS
        )

    @staticmethod
    def get_credentials():
        creds = Credentials.from_service_account_file(
           Config.CREDS_GSHEETS,
            scopes=GoogleDrive.SCOPES
        )
        return creds

    def gdrive_client_auth(self):
        client = build(
            serviceName="drive",
            version="v3",
            credentials=self.creds,
            cache_discovery=False
        )
        return client

    def connect_to_workbook(self, workbook, o_type):
        if o_type == "key":
            return self.gsheets_client.open_by_key(workbook)
        else:
            return self.gsheets_client.open(workbook)


class GSheetsWorkbook(GoogleDrive):

    def __init__(self, workbook, o_type="name"):
        """
            :arg workbook: Workbook name or workbook id
            :arg o_type:
                - open workbook by name --> name
                - open wokrbook by id --> key
        """
        super().__init__()
        self.workbook = self.connect_to_workbook(workbook, o_type)

    def get_worksheet(self, sheetname):
        return self.workbook.worksheet(sheetname)

    def create_worksheet(self, worksheet_name, num_rows, num_cols):
        worksheet = self.workbook.add_worksheet(
            title=worksheet_name, rows=num_rows, cols=num_cols
        )
        return worksheet

    def delete_worksheet(self, worksheet_name):
        self.workbook.del_worksheet(self.workbook.worksheet(worksheet_name))

    @staticmethod
    def check_worksheet_size(worksheet, required_rows):
        current_rows = worksheet.row_count
        if required_rows > current_rows:
            worksheet.resize(required_rows)

    def get_spreadsheet_data(self, sheetname, return_type="dataframe", has_filters=True):
        w = self.get_worksheet(sheetname)
        if has_filters:
            w.clear_basic_filter()
        data = w.get_all_values()
        if return_type == "dataframe":
            return pd.DataFrame(data[1:], columns=data[0])
        else:
            return data

    @staticmethod
    def find_row_by_value( worksheet, value):
        try:
            cell = worksheet.find(value)
            return cell.row
        except:
            return -1

    def update_sheets_data(self, sheetname, start_cell, body, params=None):
        """
        Update google sheets data given values to update
        If values already exist, it does nothing, otherwise, it appends new rows
            sh --> Spreadsheet: Google sheets workbook connection
            sheetname --> str: Name the sheet where data has to be uploaded
            start_cell --> str: First cell from where data will be uploaded [format -> A2, for cell(2,2)]
            body --> {"values": values, "majorDimension": "COLUMNS"}
                values: list: Values to be updates as list of lists, where each list represents a row
                other params: https://developers.google.com/sheets/api/reference/rest/v4/spreadsheets.values#ValueRange
        """
        start_range_str = f"{sheetname}!{start_cell}"
        self.workbook.values_update(
            range=start_range_str,
            params={'valueInputOption': 'RAW'} if params is None else params,
            body=body
        )

    def write_dataframe_to_new_worksheet(self, sheetname, df):
        self.create_worksheet(
            worksheet_name=sheetname,
            num_rows=len(df)+1,
            num_cols=len(df.columns)
        )
        set_with_dataframe(self.get_worksheet(sheetname), df)

    @staticmethod
    def next_available_row(worksheet, cols_to_sample=2):
        """ Looks for empty row based on values appearing in 1st N columns """
        cols = worksheet.range(1, 1, worksheet.row_count, cols_to_sample)
        return max([cell.row for cell in cols if cell.value]) + 1

    def clear_range(self, str_range):
        self.workbook.values_clear(str_range)


class Drive(GoogleDrive):

    def __init__(self):
        super().__init__()
        self.gdrive_client = self.gdrive_client_auth()

    def find_file_in_folder_by_name(self, folder_id, file_name):
        file_id = None
        folder_data = self.gdrive_client.files().list(
            q=f"'{folder_id}' in parents",
            pageSize=10,
            driveId='0AK9umGAgsOj2Uk9PVA',
            corpora='drive',
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            fields="nextPageToken, files(id, name)"
        ).execute()
        items = folder_data.get('files', [])
        file = [file for file in items if file.get("name") == file_name]
        if file:
            file_id = file[0].get('id')
        return file_id

    def read_file_from_sheets(self, file_id, file_format, tmp_folder, sheet_name=0):
        df = pd.DataFrame()
        try:
            # os.mkdir(tmp_folder)
            request = self.gdrive_client.files().get_media(fileId=file_id)
            with io.FileIO(f"{tmp_folder}/tmp_file.{file_format}", 'wb') as fh:
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                    print(f"Download {int(status.progress() * 100)}%.")

                if file_format == "xlsx":
                    df = pd.read_excel(io=f"{tmp_folder}/tmp_file.{file_format}", sheet_name=sheet_name)
                elif file_format == "csv":
                    df = pd.read_excel(io=f"{tmp_folder}/tmp_file.{file_format}")
        except Exception as e:
            print(f"Impossible to download and read data from drive. File id{file_id}", e)
        finally:
            return df

    def get_file_data_from_drive(self, folder_id, file_name, tmp_folder, file_format, sheet_name=0):
        file_id = self.find_file_in_folder_by_name(
            folder_id=folder_id,
            file_name=file_name
        )
        df = self.read_file_from_sheets(
            file_id=file_id,
            file_format=file_format,
            tmp_folder=tmp_folder,
            sheet_name=sheet_name
        )
        remove_file(os.path.join(tmp_folder, f"tmp_file.{file_format}"))
        return df

    def upload_file_to_drive(self, file_name, file_path, mimetype, drive_filename=None, destination_folder=None):
        try:
            file_metadata = {
                'name': drive_filename if drive_filename is not None else file_name,  # Desired name
                'parents': [destination_folder] if destination_folder is not None else []  # Destination folder
            }

            file_to_upload = os.path.join(file_path, file_name)
            media_content = MediaFileUpload(file_to_upload, mimetype=mimetype)

            file = self.gdrive_client.files().create(
                body=file_metadata,
                media_body=media_content
            ).execute()

            if file.get("id") is not None:
                print(f"File {file_name} successfully saved to Google Drive")

        except Exception as e:
            print("Couldn't save file to Googl Drive.", e)


def register_results(workbook, core_response_items, execution_time, clients):
    if core_response_items:
        df = pd.DataFrame(core_response_items, columns=['status', 'sourceUri', 'sourceName', 'url'])
        df.insert(loc=0, column='execution_time', value=execution_time)
        df['clients'] = None
        for idx, row in df.iterrows():
            client_match_lst = list(
                filter(lambda client: row['url'] in [news.get('link') for news in client.news], clients)
            )
            row['clients'] = ','.join([client.name for client in client_match_lst])

        worksheet = workbook.get_worksheet('execution_results')
        start_row = workbook.next_available_row(worksheet, cols_to_sample=4)
        workbook.check_worksheet_size(worksheet, required_rows=start_row + len(df) - 1)
        workbook.update_sheets_data(
            sheetname='execution_results',
            start_cell=f"A{start_row}",
            body={'values': df.values.tolist()}
        )


def format_column_data_type(worksheet, cols_to_format, dataframe_cols, format_pattern, format_type="NUMBER"):
    """
    Apply column data type format in google spreadsheets
    :param worksheet: gspread worksheet object
    :param cols_to_format: list of column names to format
    :param dataframe_cols: df.columns array --> get_loc property not valid with list(df)!!
    :param format_pattern: desired output format
    :param format_type: See docs for more oprtions, default NUMBER
    :return:
    """
    for colname in cols_to_format:
        col_idx = dataframe_cols.get_loc(colname)
        col_letter = get_column_letter(col_idx + 1)
        worksheet.format(
            f'{col_letter}:{col_letter}',
            {"numberFormat": {"type": format_type, "pattern": format_pattern}}
        )